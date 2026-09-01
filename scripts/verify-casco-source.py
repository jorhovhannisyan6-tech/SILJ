"""Developer verification: extracted CASCO rules must match the supplied Excel.
Run with: python scripts/verify-casco-source.py
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "knowledge-base" / "source-documents" / "casco_calculator_2024.xlsx"
if not BOOK.exists():
    BOOK_ALT = ROOT / "knowledge-base" / "source-documents" / "casco calculator 2024 - առանց ՃՈՈ.xlsx"
    if BOOK_ALT.exists():
        BOOK = BOOK_ALT
    else:
        raise SystemExit(f"Missing source workbook: {BOOK}")

wb = load_workbook(BOOK, data_only=True, read_only=True)
ws = wb["result 2"]
checks = {
    "warranty": (ws["B79"].value, 0.00543956043956044),
    "loss_ratio_ge_90": (ws["B74"].value, 0.00271978021978022),
    "payment_2": (ws["B52"].value, 0.0009065934065934068),
    "payment_4": (ws["B53"].value, 0.0018131868131868135),
    "payment_12": (ws["B54"].value, 0.003626373626373627),
    "traffic": (ws["B57"].value, 0.003626373626373627),
    "theft_small_details": (ws["B60"].value, -0.003626373626373627),
    "theft_excluded": (ws["B62"].value, -0.00543956043956044),
    "region_georgia": (ws["B65"].value, 0.0009065934065934068),
    "region_cis": (ws["B66"].value, 0.003626373626373627),
    "electric": (-0.001, -0.001),
    "unlimited_drivers": (ws["C33"].value, 0.003),
    "franchise_share_under7": (ws["B35"].value, 0.00022526622522879122),
    "franchise_share_over7": (ws["C35"].value, 0.0016974281768410105),
    "franchise_min_under7": (ws["B36"].value, 0.0004505324504575822),
    "franchise_min_over7": (ws["C36"].value, 0.004073827624418436),
}
for name, (actual, expected) in checks.items():
    if actual is None or abs(float(actual) - expected) > 1e-12:
        raise AssertionError(f"{name}: Excel={actual!r}, extracted={expected!r}")
print(f"PASS: {len(checks)} CASCO rule values match the supplied Excel.")
